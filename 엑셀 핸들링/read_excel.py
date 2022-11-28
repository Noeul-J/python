# 엑셀 읽기 위해 필요한 모듈
from openpyxl import load_workbook

filepath = "C:\RPA\\test.xlsx"

# Load_workbook으로 Load하기
wb = load_workbook(filepath)
ws = wb.active

# max_row 읽기
max_row = ws.max_row

# max_column 읽기
max_column = ws.max_column

# max row. col 을 토대로 읽어들이기
for i in range(1, max_row+1):
    for j in range(1, max_column+1):
        #get particular cell value
        print(ws.cell(row=i, column=j).value, end = '|')
    print('\n')