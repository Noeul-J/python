import openpyxl

#워크북 생성
wb = openpyxl.Workbook()

#시트 활성화
ws = wb.worksheets[0] #wb.active 로 써도 됨

#시트 이름바꾸기
ws.title = "worksheet"

#새로운 시트 만들기
ws2 = wb.create_sheet('worksheet2')

#두번째 worksheet2에 'B2'에 'test' 쓰기
ws2['B2'] = 'test'

#워크북 filename으로 저장(없으면 새로 생성)
filename = 'C:\RPA\\test.xlsx'
wb.save(filename)

wb = openpyxl.Workbook()
ws = wb.worksheets[0]
ws.title = "worksheet"