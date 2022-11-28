import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.worksheets[0]
ws.title = "worksheet"

#가운데 정렬
align_center = Alignment(horizontal='center', vertical='center')

#글씨체 굵게
font_bold = Font(size=12, bold=True, color='000000') # 000000:black

#셀 색깔 채우기
fill_blue = PatternFill('solid', fgColor='819FF7')

#테두리 선넣기
thin_border = Border(left=Side(border_style='thin', color='000000'),
                     right=Side(border_style='thin', color='000000'),
                     top=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000'))

#cell에 적용
ws['C1'].alignment = align_center
ws['C1'].font = font_bold
ws['C1'].fill = fill_blue
ws['C1'].border = thin_border
ws['C1'].value = 'text'

#범위 지정
my_range = ws['A1':'B2']

for row in my_range:
    for cell in row:
        cell.fill = fill_blue
        cell.border = thin_border
        cell.value = 'range'
        
#worksheet.cell로 접근하기
for c in range(3):
    ws.cell(row=1, column=c+1).value = "text"
    ws.cell(row=1, column=c+1).alignment = align_center

#기존에 써있는 마지막 행의 다음행에 추가됨
ws.append(['1','2','3'])

filename = 'C:\RPA\\test.xlsx'
wb.save(filename)