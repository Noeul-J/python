import openpyxl

wb = openpyxl.Workbook()
ws = wb.worksheets[0]
ws.title = "worksheet"

# 행높이, 열너비 바꾸기
ws.row_dimensions[1].height = 30
ws.column_dimensions['A'].height = 30

# for 문으로 행높이, 열너비 바꾸기
for row in range(1,14):
    ws.row_dimensions[row].height = 30

for col in range(65,70):
    ws.column_dimensions[chr(col)].width = 20

# 병합하기
ws.merge_cells("C6:D6")
ws.unmerge_cells("C6:D6")

# 날짜 표기
import datetime
ws['A1'] = datetime.datetime.now()
ws['A2'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
ws['A3'] = datetime.datetime.now().strftime("%Y년 %m월 %d일".encode('unicode-escape').decode()).encode().decode('unicode-escape')

filename = 'C:\RPA\\test.xlsx'
wb.save(filename)