import pandas as pd                 # pip install pandas
import difflib
import sys
from openpyxl import load_workbook  # pip install openpyxl
from openpyxl.styles import PatternFill, Font


def similarity_check_and_write(file_name):
    try:
        # 엑셀 파일 읽기
        ocr = pd.read_excel(file_name, sheet_name='OCR_결과')
        answer = pd.read_excel(file_name, sheet_name='정답지_결과')

        # 리스트 변수 초기화
        similar_row = []
        similar_result = []

        # 셀 비교해서 유사도 측정
        for row in range(0, len(ocr)):
            for col in ocr.columns:
                ocr_data = ocr[col][row]
                answer_data = answer[col][row]

                # 품목명, 이미지명 열 제외
                if col == "품목명" or col == "이미지명":
                    similar_row.append(ocr_data)
                    continue

                # 유사도 체크
                if pd.isna(ocr_data) and pd.isna(answer_data):
                    similar = ''
                elif ocr_data == answer_data:
                    similar = str(100.0) + "%"
                else:
                    ocr_bytes = bytes(str(ocr_data), 'utf-8')
                    answer_bytes = bytes(str(answer_data), 'utf-8')
                    ocr_bytes_list = list(ocr_bytes)
                    answer_bytes_list = list(answer_bytes)

                    sm = difflib.SequenceMatcher(lambda x: x == " ", answer_bytes_list, ocr_bytes_list)
                    similar = sm.ratio() * 100
                    similar = str(round(similar, 2))+"%"

                similar_row.append(similar)

                if col == '생산자':
                    similar_result.append(similar_row)
                    similar_row = []

        # 유사도 DataFrame 설정
        similar_df = pd.DataFrame(similar_result)
        similar_df.columns = ocr.columns
        print(similar_df)

        # 엑셀 파일에 입력
        with pd.ExcelWriter(file_name, if_sheet_exists='overlay', mode='a', engine='openpyxl') as writer:
            similar_df.to_excel(writer, sheet_name="유사도", header=True, index=False, startcol=0, startrow=0)

        return True
    except Exception as e:
        print(f"유사도 추출 중 오류 발생:\n{e}")
        return e


def check_for_no100(file_name):
    try:
        # 서식 변수 설정
        font_default = Font(size=12, bold=False, color='000000')
        font_red = Font(size=12, bold=False, color='FF0000')                  # 글씨 빨강
        # fill_colors = PatternFill('solid', fgColor='ffb399')                # 셀 색깔 채우기

        # 결과 데이터 읽기
        wb = load_workbook(file_name)
        ws = wb['유사도']
        similar_data = pd.read_excel(file_name, sheet_name='유사도')
        print(similar_data)

        # 서식 값 초기화
        for row in range(2, len(similar_data)):
            for col in range(1, len(similar_data.columns)):
                ws.cell(row=row, column=col).font = font_default

        # # 100% 가 아닌 데이터 빨간 글씨로 변경
        for row in range(0, len(similar_data)):
            col_num = 3
            for col in similar_data.columns:
                data = similar_data[col][row]
                if col == '품목명' or col == '이미지명':
                    continue
                if pd.isna(data):
                    col_num += 1
                    continue

                if data != "100.0%":
                    ws.cell(row=row+2, column=col_num).font = font_red

                col_num += 1

        wb.save(file_name)
        return True
    except Exception as e:
        print(f"유사도 서식 변경 중 오류 발생:\n{e}")
        return e


if __name__ == '__main__':
    #excel_file_name = 'ocr_result.xlsx'
    excel_file_name = sys.argv[1]
    smchk = similarity_check_and_write(excel_file_name)
    if smchk != True:
        fail_msg = "Fail:"+str(smchk)
        print("<peon>")
        print(fail_msg)
        print("</peon>")
        sys.exit()

    chkno100 = check_for_no100(excel_file_name)
    if chkno100:
        print("<peon>")
        print(chkno100)
        print("</peon>")
    else:
        fail_msg = "Fail:"+str(chkno100)
        print("<peon>")
        print(fail_msg)
        print("</peon>")

