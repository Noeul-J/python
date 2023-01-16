from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np

filename = 'C:\RPA\현금영수증.xlsx'
sheetName = '2146'
startCol = 'A'

outputFileName = filename[0:filename.rfind('.')] + '-new.xlsx'

wb = load_workbook(filename)
#wb.get_sheet_names()

ws = wb[sheetName]

max_row = ws.max_row
print(max_row)

