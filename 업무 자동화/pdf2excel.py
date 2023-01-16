import pdfplumber
import pandas as pd
import openpyxl as xls
import os


def pdf2table(path,fname) -> str:
    pdf_fname = os.path.join(path, fname + '.pdf')
    pdf = pdfplumber.open(pdf_fname)    # pdf 열기
    page = pdf.pages[0]             # 첫번째 페이지

    tables = page.extract_tables()  # 첫번째 페이지의 모든 테이블 추출
    df = pd.DataFrame(tables[0])
    print(df.head())

    result_fname = os.path.join(path, fname + '.xlsx')  # 데이터를 엑셀로 저장
    print(result_fname)
    df.to_excel(result_fname)

    pdf.close()                     # pdf 닫기

    return result_fname


def cell_merge(excel_fname):
    # 엑셀 첫 행, 첫 열 삭제
    wb = xls.load_workbook(excel_fname)
    ws = wb.active

    ws.delete_rows(1)
    ws.delete_cols(1)

    wb.save(excel_fname)

    # 엑셀 데이터 읽기
    df = pd.read_excel(excel_fname)
    print(df)
    print(ws.max_column)

    # 헤더 행 열 위치 가져오기
    hd_col_list = []
    for hd_col_idx in range(1, ws.max_column+1):
        if ws.cell(row=1, column=hd_col_idx).value is None:
            ws.merge_cells(start_row=1, start_column=hd_col_idx-1, end_row=1, end_column=hd_col_idx)
        else:
            hd_col_list.append(hd_col_idx)

    print(hd_col_list)
    wb.save(excel_fname)

    # 세로 행 위치 가져오기
    hd_row_list = []
    for hd_row_idx in range(1, ws.max_row+1):
        if ws.cell(row=hd_row_idx, column=1).value is not None:
            hd_row_list.append(hd_row_idx)

    print(hd_row_list)


    # 가로 셀 병합
    for row_idx in range(1, ws.max_row+1):
        for col_idx in range(1, ws.max_column+1):
            if col_idx not in hd_col_list and ws.cell(row=row_idx, column=col_idx).value is None:
                ws.merge_cells(start_row=row_idx, start_column=col_idx-1, end_row=row_idx, end_column=col_idx)
                wb.save(excel_fname)

    # 세로 셀 병합
    for row_idx in range(1, ws.max_row+1):
        for col_idx in range(1, ws.max_column + 1):
            if row_idx not in hd_row_list and ws.cell(row=row_idx, column=col_idx).value is None:
                ws.merge_cells(start_row=row_idx-1, start_column=col_idx, end_row=row_idx, end_column=col_idx)
                wb.save(excel_fname)
            # # 해당 셀 값과 윗줄 셀 값을 가져옴
            # val = ws.cell(row=row_index, column=col_index).value
            # if row_index == 1:
            #     up_cell_value = '첫줄'
            # else:
            #     up_cell_value = ws.cell(row=row_index-1, column=col_index).value
            #
            # # 해당 셀 값과 윗줄 셀 값이 모두 None 이면 셀 가로 병합
            # if val is None and up_cell_value is None:
            #     print(val)
            #     print(str(row_index) + " , " + str(col_index))
            #     ws.merge_cells(start_row=row_index, start_column=col_index-1, end_row=row_index, end_column=col_index)
            #     wb.save(excel_fname)

    wb.save(excel_fname)


if __name__ == '__main__':
    # excel_fname_from_wd = pdf2table('C:\\RPA\\TEST\\정보공개서\\', 'test')
    excel_fname_from_wd = pdf2table('C:\\RPA\\TEST\\', 'GHTEST')
    cell_merge(excel_fname_from_wd)
    # print(text_from_pdf)
